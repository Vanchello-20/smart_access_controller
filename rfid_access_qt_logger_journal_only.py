import os
import re
import sys
import time
from dataclasses import dataclass
from datetime import datetime

try:
    import serial
    from serial.tools import list_ports
except ImportError:
    print("Не установлен pyserial. Выполни: py -m pip install pyserial")
    sys.exit(1)

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Не установлен openpyxl. Выполни: py -m pip install openpyxl")
    sys.exit(1)

try:
    from PyQt6.QtCore import QThread, pyqtSignal, Qt, QTimer
    from PyQt6.QtGui import QColor, QFont, QIcon
    from PyQt6.QtWidgets import (
        QApplication, QComboBox, QFormLayout, QFrame, QGridLayout, QGroupBox,
        QHBoxLayout, QHeaderView, QLabel, QLineEdit, QMainWindow, QMessageBox,
        QPushButton, QSpinBox, QStatusBar, QTableWidget, QTableWidgetItem,
        QVBoxLayout, QWidget
    )
except ImportError:
    print("Не установлен PyQt6. Выполни: py -m pip install PyQt6")
    sys.exit(1)

APP_TITLE = "Access Control System"
APP_ICON_FILE = "KeyChainAccess_37052.ico"
DEFAULT_BAUDRATE = 115200
WORKBOOK_PATH = "access_control_logger.xlsx"
LOG_SHEET = "Журнал"
DEFAULT_UID_CAPACITY = 20

LOG_HEADERS = [
    "Дата", "Время", "UID", "Фамилия", "Имя", "Кабинет",
    "Статус", "Сырой пакет", "Время записи ПК"
]
CARD_HEADERS = ["UID", "Фамилия", "Имя", "Кабинет", "Разрешён", "Комментарий"]

# Excel запрещает управляющие символы ASCII 0..31, кроме обычных переводов строк.
# Arduino при перезапуске/первом подключении иногда успевает выплюнуть мусор в Serial,
# поэтому вход надо чистить до таблицы и до .xlsx, а не надеяться на чудо.
EXCEL_ILLEGAL_RE = re.compile(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]")
PRINTABLE_PACKET_RE = re.compile(r"[^\x20-\x7EА-Яа-яЁё№|:;.,_+\-\/\\()\[\]{}@#%&*!?=<> ]")


def app_resource_path(filename: str) -> str:
    """Возвращает путь к файлу рядом со скриптом или собранным exe."""
    candidates = []

    if getattr(sys, "frozen", False):
        candidates.append(os.path.dirname(sys.executable))

    candidates.extend([
        os.path.dirname(os.path.abspath(__file__)),
        os.getcwd(),
    ])

    for base_dir in candidates:
        path = os.path.join(base_dir, filename)
        if os.path.exists(path):
            return path

    return os.path.join(candidates[0], filename) if candidates else filename


def load_app_icon() -> QIcon:
    icon_path = app_resource_path(APP_ICON_FILE)
    if os.path.exists(icon_path):
        return QIcon(icon_path)
    return QIcon()


def clean_excel_value(value) -> str:
    """Удаляет символы, из-за которых openpyxl/Excel падают при записи."""
    if value is None:
        return ""
    return EXCEL_ILLEGAL_RE.sub("", str(value)).strip()


def normalize_serial_line(line: str) -> str:
    """Чистит строку из COM-порта от мусора запуска и управляющих байтов."""
    line = clean_excel_value(line)
    line = line.replace("\r", "").replace("\n", "").strip()
    # Оставляем ожидаемые символы протокола и русские буквы.
    line = PRINTABLE_PACKET_RE.sub("", line).strip()
    return line


def is_expected_packet(line: str) -> bool:
    """Пропускает только нормальные пакеты протокола, а стартовый мусор Arduino режет."""
    if not line:
        return False
    return (
        line.startswith("LOG|")
        or line.startswith("STATUS|")
        or line.startswith("GRANTED|")
        or line.startswith("DENIED|")
        or line.startswith("INFO|")
        or line.startswith("ERROR|")
    )


@dataclass
class LogEvent:
    date: str
    time: str
    uid: str
    surname: str
    name: str
    room: str
    status: str
    raw: str
    pc_time: str

    def row(self):
        return [
            clean_excel_value(self.date),
            clean_excel_value(self.time),
            clean_excel_value(self.uid),
            clean_excel_value(self.surname),
            clean_excel_value(self.name),
            clean_excel_value(self.room),
            clean_excel_value(self.status),
            clean_excel_value(self.raw),
            clean_excel_value(self.pc_time),
        ]


class ExcelLogger:
    def __init__(self, path=WORKBOOK_PATH):
        self.path = path
        self.ensure_workbook()

    def ensure_workbook(self):
        if os.path.exists(self.path):
            wb = load_workbook(self.path)
        else:
            wb = Workbook()

        if LOG_SHEET in wb.sheetnames:
            ws = wb[LOG_SHEET]
        else:
            ws = wb.active
            ws.title = LOG_SHEET

        for sheet_name in list(wb.sheetnames):
            if sheet_name != LOG_SHEET:
                del wb[sheet_name]

        self._prepare_log(wb)
        wb.save(self.path)

    def _prepare_log(self, wb):
        ws = wb[LOG_SHEET]

        for i, header in enumerate(LOG_HEADERS, 1):
            ws.cell(1, i).value = header

        self._style_header(ws, "DDEBFF")

        widths = [14, 12, 24, 18, 18, 12, 18, 64, 22]
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:I{max(ws.max_row, 2)}"

        for row in range(2, ws.max_row + 1):
            for col in range(1, len(LOG_HEADERS) + 1):
                ws.cell(row, col).alignment = Alignment(
                    vertical="center",
                    wrap_text=(col == 8)
                )

    def _style_header(self, ws, fill):
        thin = Side(style="thin", color="D9E2F3")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="1F2937")
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(bottom=thin)
        ws.row_dimensions[1].height = 22

    def append_event(self, event: LogEvent):
        wb = load_workbook(self.path)

        for sheet_name in list(wb.sheetnames):
            if sheet_name != LOG_SHEET:
                del wb[sheet_name]

        self._prepare_log(wb)
        ws = wb[LOG_SHEET]
        ws.append(event.row())

        row = ws.max_row
        status_cell = ws.cell(row, 7)
        status = event.status.upper()

        if status in ("GRANTED", "ACCESS_GRANTED", "OK", "CARD_WRITTEN", "UID_ADDED", "CARD_ERASED", "UID_REMOVED", "UIDS_CLEARED"):
            status_cell.fill = PatternFill("solid", fgColor="E2F0D9")
            status_cell.font = Font(color="375623", bold=True)
        elif status in ("DENIED", "ACCESS_DENIED", "ERROR", "UID_ADD_FAILED", "CARD_WRITE_FAILED", "CARD_ERASE_FAILED"):
            status_cell.fill = PatternFill("solid", fgColor="FCE4D6")
            status_cell.font = Font(color="9C0006", bold=True)
        else:
            status_cell.fill = PatternFill("solid", fgColor="F2F2F2")

        for col in range(1, len(LOG_HEADERS) + 1):
            ws.cell(row, col).alignment = Alignment(
                vertical="center",
                wrap_text=(col == 8)
            )

        ws.auto_filter.ref = f"A1:I{ws.max_row}"
        wb.save(self.path)

    def clear_log_data(self):
        wb = load_workbook(self.path)

        for sheet_name in list(wb.sheetnames):
            if sheet_name != LOG_SHEET:
                del wb[sheet_name]

        self._prepare_log(wb)
        ws = wb[LOG_SHEET]
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(LOG_HEADERS)):
            for cell in row:
                cell.fill = PatternFill()
                cell.font = Font(color="000000", bold=False)

        ws.auto_filter.ref = "A1:I1"
        wb.save(self.path)

def parse_line(line: str) -> LogEvent:
    now = datetime.now()
    date = now.strftime("%Y-%m-%d")
    clock = now.strftime("%H:%M:%S")
    pc_time = now.strftime("%Y-%m-%d %H:%M:%S")

    line = normalize_serial_line(line)
    parts = [clean_excel_value(part) for part in line.split("|")]

    # Формат текущей прошивки Arduino:
    # LOG|EVENT|UID|SURNAME|NAME|OFFICE|SOURCE|NOTE
    # Пример:
    # LOG|ACCESS_GRANTED|B3 45 0A 30|Ivan|Konin|2|CARD|
    if len(parts) >= 7 and parts[0] == "LOG":
        event_type = parts[1]
        uid = parts[2]
        surname = parts[3]
        name = parts[4]
        room = parts[5]
        return LogEvent(date, clock, uid, surname, name, room, event_type, line, pc_time)

    # Служебные сообщения Arduino, например: STATUS|MODE|ACCESS
    if len(parts) >= 2 and parts[0] == "STATUS":
        return LogEvent(date, clock, "", "", "", "", "INFO", line, pc_time)

    # Старый/резервный формат, если когда-нибудь понадобится:
    # GRANTED|UID|SURNAME|NAME|OFFICE|...
    if len(parts) >= 5 and parts[0] in ("GRANTED", "DENIED"):
        uid = parts[1]
        surname = parts[2]
        name = parts[3]
        room = parts[4]
        return LogEvent(date, clock, uid, surname, name, room, parts[0], line, pc_time)

    if parts and parts[0] in ("INFO", "ERROR"):
        status = parts[0]
        text = "|".join(parts[1:]) if len(parts) > 1 else line
        return LogEvent(date, clock, "", "", "", "", status, text, pc_time)

    return LogEvent(date, clock, "", "", "", "", "INFO", line, pc_time)


def safe_decode(raw: bytes) -> str:
    for enc in ("utf-8", "cp1251", "latin1"):
        try:
            return raw.decode(enc).strip()
        except UnicodeDecodeError:
            pass
    return raw.decode("utf-8", errors="replace").strip()


class SerialWorker(QThread):
    line_received = pyqtSignal(str)
    status_changed = pyqtSignal(str)
    error = pyqtSignal(str)
    connected = pyqtSignal(bool)

    def __init__(self, port: str, baudrate: int):
        super().__init__()
        self.port = port
        self.baudrate = baudrate
        self._running = True
        self.ser = None

    def run(self):
        try:
            self.ser = serial.Serial(self.port, self.baudrate, timeout=0.15)
            time.sleep(2)
            self.connected.emit(True)
            self.status_changed.emit(f"Подключено: {self.port}, {self.baudrate}")
            while self._running:
                if self.ser.in_waiting:
                    raw = self.ser.readline()
                    line = normalize_serial_line(safe_decode(raw))
                    if is_expected_packet(line):
                        self.line_received.emit(line)
                    elif line:
                        self.status_changed.emit(f"Пропущен мусор из COM: {line[:40]}")
                else:
                    self.msleep(40)
        except serial.SerialException as exc:
            self.error.emit(f"Не удалось открыть порт {self.port}: {exc}")
        finally:
            try:
                if self.ser and self.ser.is_open:
                    self.ser.close()
            except Exception:
                pass
            self.connected.emit(False)
            self.status_changed.emit("Отключено")

    def stop(self):
        self._running = False
        self.wait(1200)

    def send(self, text: str):
        if self.ser and self.ser.is_open:
            self.ser.write((text.strip() + "\n").encode("utf-8"))
            self.status_changed.emit(f"Отправлено: {text.strip()}")
        else:
            self.error.emit("Порт не открыт")


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.logger = ExcelLogger(WORKBOOK_PATH)
        self.worker = None
        self.events_total = 0
        self.granted_total = 0
        self.denied_total = 0
        self.arduino_mode = "неизвестно"
        self.uid_count = None
        self.uid_capacity = DEFAULT_UID_CAPACITY
        self.setWindowTitle(APP_TITLE)
        self.setWindowIcon(load_app_icon())
        self.resize(1180, 760)
        self._build_ui()
        #self._apply_style()
        self.refresh_ports()
        self.statusBar().showMessage(f"Excel-журнал с одним листом: {os.path.abspath(WORKBOOK_PATH)}")

    def _build_ui(self):
        root = QWidget()
        self.setCentralWidget(root)
        layout = QVBoxLayout(root)
        layout.setContentsMargins(18, 18, 18, 14)
        layout.setSpacing(12)

        '''title = QLabel("RFID Access Control")
        title.setObjectName("Title")
        subtitle = QLabel("Живой журнал проходов, управление режимами Arduino и сохранение в Excel без VBA/MSComm. Мир стал чуть менее абсурдным.")
        subtitle.setObjectName("Subtitle")
        subtitle.setWordWrap(True)
        layout.addWidget(title)
        layout.addWidget(subtitle)'''

        top_grid = QGridLayout()
        top_grid.setSpacing(12)
        layout.addLayout(top_grid)

        connection = QGroupBox("Подключение")
        conn_layout = QGridLayout(connection)
        self.port_box = QComboBox()
        self.baud_spin = QSpinBox()
        self.baud_spin.setRange(1200, 115200)
        self.baud_spin.setValue(DEFAULT_BAUDRATE)
        self.baud_spin.setSingleStep(1200)
        self.refresh_btn = QPushButton("Обновить порты")
        self.connect_btn = QPushButton("Подключиться")
        self.disconnect_btn = QPushButton("Отключиться")
        self.disconnect_btn.setEnabled(False)
        conn_layout.addWidget(QLabel("COM-порт"), 0, 0)
        conn_layout.addWidget(self.port_box, 0, 1, 1, 2)
        conn_layout.addWidget(QLabel("Скорость"), 1, 0)
        conn_layout.addWidget(self.baud_spin, 1, 1, 1, 2)
        conn_layout.addWidget(self.refresh_btn, 2, 0)
        conn_layout.addWidget(self.connect_btn, 2, 1)
        conn_layout.addWidget(self.disconnect_btn, 2, 2)
        top_grid.addWidget(connection, 0, 0)

        modes = QGroupBox("Режимы и дверь")
        modes_layout = QGridLayout(modes)
        self.access_btn = QPushButton("ACCESS")
        self.uid_btn = QPushButton("UID: добавить/удалить")
        self.erase_card_btn = QPushButton("Стереть карту")
        self.open_btn = QPushButton("Открыть дверь")
        modes_layout.addWidget(self.access_btn, 0, 0)
        modes_layout.addWidget(self.uid_btn, 0, 1)
        modes_layout.addWidget(self.erase_card_btn, 1, 0)
        modes_layout.addWidget(self.open_btn, 1, 1)
        top_grid.addWidget(modes, 0, 1)

        write_box = QGroupBox("Запись данных на карту")
        form = QFormLayout(write_box)
        self.surname_edit = QLineEdit()
        self.name_edit = QLineEdit()
        self.room_edit = QLineEdit()
        self.send_write_btn = QPushButton("Отправить WRITE и приложить карту")
        form.addRow("Фамилия", self.surname_edit)
        form.addRow("Имя", self.name_edit)
        form.addRow("Кабинет", self.room_edit)
        form.addRow(self.send_write_btn)
        top_grid.addWidget(write_box, 0, 2)

        stats = QFrame()
        stats.setObjectName("Stats")
        stats_layout = QHBoxLayout(stats)
        self.total_label = QLabel("Всего: 0")
        self.granted_label = QLabel("Разрешено: 0")
        self.denied_label = QLabel("Отказано: 0")
        self.mode_label = QLabel("Режим Arduino: неизвестно")
        self.uid_memory_label = QLabel(f"Карты: ? / {self.uid_capacity}, осталось: ?")
        self.last_label = QLabel("Последнее событие: —")
        for w in [self.total_label, self.granted_label, self.denied_label, self.mode_label, self.uid_memory_label, self.last_label]:
            w.setObjectName("StatLabel")
            stats_layout.addWidget(w)
        stats_layout.addStretch()
        layout.addWidget(stats)

        self.table = QTableWidget(0, len(LOG_HEADERS))
        self.table.setHorizontalHeaderLabels(LOG_HEADERS)
        self.table.verticalHeader().setVisible(False)
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        for i, width in enumerate([95, 80, 155, 120, 120, 90, 120, 340, 150]):
            self.table.setColumnWidth(i, width)
        layout.addWidget(self.table, stretch=1)

        bottom = QHBoxLayout()
        self.clear_btn = QPushButton("Очистить экран")
        self.clear_excel_btn = QPushButton("Очистить Excel-файл")
        self.clear_arduino_cards_btn = QPushButton("Удалить все карты из Arduino")
        self.open_excel_btn = QPushButton("Открыть Excel-файл")
        self.file_label = QLabel(os.path.abspath(WORKBOOK_PATH))
        self.file_label.setTextInteractionFlags(Qt.TextInteractionFlag.TextSelectableByMouse)
        bottom.addWidget(self.clear_btn)
        bottom.addWidget(self.clear_excel_btn)
        bottom.addWidget(self.clear_arduino_cards_btn)
        bottom.addWidget(self.open_excel_btn)
        bottom.addWidget(self.file_label, stretch=1)
        layout.addLayout(bottom)

        self.setStatusBar(QStatusBar())

        self.refresh_btn.clicked.connect(self.refresh_ports)
        self.connect_btn.clicked.connect(self.connect_serial)
        self.disconnect_btn.clicked.connect(self.disconnect_serial)
        self.access_btn.clicked.connect(lambda: self.send_command("MODE|ACCESS"))
        self.uid_btn.clicked.connect(lambda: self.send_command("MODE|UID"))
        self.erase_card_btn.clicked.connect(self.prepare_card_erase)
        self.open_btn.clicked.connect(lambda: self.send_command("OPEN"))
        self.send_write_btn.clicked.connect(self.prepare_card_write)
        self.clear_btn.clicked.connect(lambda: self.table.setRowCount(0))
        self.clear_excel_btn.clicked.connect(self.clear_excel_file)
        self.clear_arduino_cards_btn.clicked.connect(self.clear_arduino_cards)
        self.open_excel_btn.clicked.connect(self.open_excel)
        self._set_controls_enabled(False)

    def _apply_style(self):
        self.setStyleSheet("""
            QMainWindow { background: #F5F7FB; }
            QLabel#Title { font-size: 26px; font-weight: 700; color: #1F2937; }
            QLabel#Subtitle { color: #4B5563; font-size: 12px; }
            QGroupBox { background: white; border: 1px solid #D8E0EF; border-radius: 14px; margin-top: 12px; padding: 12px; font-weight: 700; color: #1F2937; }
            QGroupBox::title { subcontrol-origin: margin; left: 14px; padding: 0 6px; }
            QPushButton { background: #E8F0FE; border: 1px solid #C9D8F5; border-radius: 10px; padding: 9px 12px; color: #1F2937; font-weight: 600; }
            QPushButton:hover { background: #DCEAFE; }
            QPushButton:pressed { background: #C9DAFA; }
            QPushButton:disabled { background: #EEF1F6; color: #9AA3B2; border-color: #E1E5EC; }
            QComboBox, QSpinBox, QLineEdit { background: white; border: 1px solid #CBD5E1; border-radius: 8px; padding: 7px; }
            QFrame#Stats { background: white; border: 1px solid #D8E0EF; border-radius: 14px; }
            QLabel#StatLabel { padding: 10px 12px; font-weight: 700; color: #334155; }
            QTableWidget { background: white; alternate-background-color: #F8FAFC; border: 1px solid #D8E0EF; border-radius: 12px; gridline-color: #E5EAF3; }
            QHeaderView::section { background: #DDEBFF; color: #1F2937; padding: 8px; border: none; border-right: 1px solid #C9D8F5; font-weight: 700; }
        """)

    def refresh_ports(self):
        current = self.port_box.currentText()
        self.port_box.clear()
        ports = list(list_ports.comports())
        for port in ports:
            self.port_box.addItem(f"{port.device} — {port.description}", port.device)
        if current:
            index = self.port_box.findText(current)
            if index >= 0:
                self.port_box.setCurrentIndex(index)
        if not ports:
            self.statusBar().showMessage("COM-порты не найдены. Проверь кабель и драйвер Arduino.")

    def connect_serial(self):
        port = self.port_box.currentData()
        if not port:
            QMessageBox.warning(self, "Нет COM-порта", "Arduino не найден. Проверь кабель, порт и драйвер.")
            return
        self.worker = SerialWorker(port, self.baud_spin.value())
        self.worker.line_received.connect(self.handle_line)
        self.worker.status_changed.connect(self.statusBar().showMessage)
        self.worker.error.connect(self.show_error)
        self.worker.connected.connect(self.on_connection_changed)
        self.worker.start()

    def disconnect_serial(self):
        if self.worker:
            self.worker.stop()
            self.worker = None
        self.on_connection_changed(False)

    def on_connection_changed(self, connected: bool):
        self.connect_btn.setEnabled(not connected)
        self.disconnect_btn.setEnabled(connected)
        self.port_box.setEnabled(not connected)
        self.baud_spin.setEnabled(not connected)
        self.refresh_btn.setEnabled(not connected)
        self._set_controls_enabled(connected)
        if not connected:
            self.arduino_mode = "неизвестно"
            self.uid_count = None
            self.update_arduino_status_labels()
        else:
            QTimer.singleShot(300, lambda: self.send_command("STATUS"))

    def _set_controls_enabled(self, enabled: bool):
        for btn in [self.access_btn, self.uid_btn, self.erase_card_btn, self.open_btn, self.send_write_btn, self.clear_arduino_cards_btn]:
            btn.setEnabled(enabled)

    def send_command(self, command: str):
        if not self.worker:
            self.show_error("Сначала подключись к COM-порту")
            return
        self.worker.send(command)

    def prepare_card_write(self):
        surname = self.surname_edit.text().strip()
        name = self.name_edit.text().strip()
        room = self.room_edit.text().strip()
        if not surname or not name or not room:
            QMessageBox.warning(self, "Пустые поля", "Заполни фамилию, имя и кабинет. Карта не гадалка.")
            return
        self.send_command("MODE|WRITE")
        QTimer.singleShot(150, lambda: self.send_command(f"WRITE|{surname}|{name}|{room}"))
        self.statusBar().showMessage("Команда записи отправлена. Теперь приложи карту к считывателю.")

    def prepare_card_erase(self):
        self.send_command("MODE|ERASE")
        self.statusBar().showMessage("Режим стирания включен. Приложи карту: данные будут стерты, UID удален из разрешенных.")

    def clear_excel_file(self):
        answer = QMessageBox.question(
            self,
            "Очистить Excel-файл",
            "Удалить все строки журнала из Excel-файла и сбросить цветные заливки?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        )
        if answer != QMessageBox.StandardButton.Yes:
            return

        try:
            self.logger.clear_log_data()
            self.statusBar().showMessage("Excel-файл очищен: строки данных и цветные заливки удалены.")
        except PermissionError:
            self.show_error("Excel-файл открыт. Закрой его, чтобы программа смогла очистить .xlsx")
        except Exception as exc:
            self.show_error(f"Ошибка очистки Excel: {exc}")

    def clear_arduino_cards(self):
        answer = QMessageBox.question(
            self,
            "Удалить все карты из Arduino",
            "Удалить из памяти Arduino все разрешенные UID? Данные на самих картах это не стирает.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        )
        if answer != QMessageBox.StandardButton.Yes:
            return

        self.send_command("CLEAR_UIDS")
        self.statusBar().showMessage("Команда удаления всех UID из Arduino отправлена.")

    def update_arduino_status_labels(self):
        mode_names = {
            "ACCESS": "ACCESS",
            "UID_MANAGE": "UID_MANAGE",
            "CARD_WRITE": "запись карты",
            "CARD_ERASE": "стирание карты",
        }
        self.mode_label.setText(f"Режим Arduino: {mode_names.get(self.arduino_mode, self.arduino_mode)}")

        if self.uid_count is None:
            self.uid_memory_label.setText(f"Карты: ? / {self.uid_capacity}, осталось: ?")
            return

        free_slots = max(self.uid_capacity - self.uid_count, 0)
        self.uid_memory_label.setText(f"Карты: {self.uid_count} / {self.uid_capacity}, осталось: {free_slots}")

    def handle_status_packet(self, line: str):
        parts = [clean_excel_value(part) for part in line.split("|")]
        if len(parts) < 3 or parts[0] != "STATUS":
            return

        kind = parts[1]
        value = parts[2]

        if kind == "MODE":
            self.arduino_mode = value
            self.update_arduino_status_labels()
        elif kind == "UID_COUNT":
            try:
                self.uid_count = int(value)
                self.update_arduino_status_labels()
            except ValueError:
                pass
        elif kind == "UID_CAPACITY":
            try:
                self.uid_capacity = int(value)
                self.update_arduino_status_labels()
            except ValueError:
                pass

    def handle_line(self, line: str):
        line = normalize_serial_line(line)
        if not is_expected_packet(line):
            self.statusBar().showMessage(f"Пропущен мусор из COM: {line[:40]}")
            return
        self.handle_status_packet(line)
        event = parse_line(line)
        self.add_event_to_table(event)
        try:
            self.logger.append_event(event)
        except PermissionError:
            self.show_error("Excel-файл открыт. Закрой его, чтобы программа смогла сохранить строку в .xlsx")
        except Exception as exc:
            self.show_error(f"Ошибка записи в Excel: {exc}")

    def add_event_to_table(self, event: LogEvent):
        row = self.table.rowCount()
        self.table.insertRow(row)
        values = event.row()
        for col, value in enumerate(values):
            item = QTableWidgetItem(str(value))
            item.setTextAlignment(Qt.AlignmentFlag.AlignVCenter | (Qt.AlignmentFlag.AlignCenter if col in (0, 1, 5, 6) else Qt.AlignmentFlag.AlignLeft))
            if col == 6:
                status = event.status.upper()
                if status in ("GRANTED", "ACCESS_GRANTED", "OK", "CARD_WRITTEN", "UID_ADDED", "CARD_ERASED", "UID_REMOVED", "UIDS_CLEARED"):
                    item.setBackground(QColor("#E2F0D9"))
                    item.setForeground(QColor("#375623"))
                    font = QFont()
                    font.setBold(True)
                    item.setFont(font)
                elif status in ("DENIED", "ACCESS_DENIED", "ERROR", "CARD_WRITE_FAILED", "UID_ADD_FAILED", "CARD_ERASE_FAILED"):
                    item.setBackground(QColor("#FCE4D6"))
                    item.setForeground(QColor("#9C0006"))
                    font = QFont()
                    font.setBold(True)
                    item.setFont(font)
            self.table.setItem(row, col, item)
        self.table.scrollToBottom()
        self.events_total += 1
        if event.status.upper() in ("GRANTED", "ACCESS_GRANTED", "OK", "CARD_WRITTEN", "UID_ADDED", "CARD_ERASED", "UID_REMOVED", "UIDS_CLEARED"):
            self.granted_total += 1
        elif event.status.upper() in ("DENIED", "ACCESS_DENIED", "CARD_WRITE_FAILED", "UID_ADD_FAILED", "CARD_ERASE_FAILED"):
            self.denied_total += 1
        self.total_label.setText(f"Всего: {self.events_total}")
        self.granted_label.setText(f"Разрешено: {self.granted_total}")
        self.denied_label.setText(f"Отказано: {self.denied_total}")
        self.last_label.setText(f"Последнее событие: {event.time} | {event.uid or event.raw[:28]}")

    def open_excel(self):
        self.logger.ensure_workbook()
        path = os.path.abspath(WORKBOOK_PATH)
        try:
            os.startfile(path)
        except AttributeError:
            QMessageBox.information(self, "Файл", path)
        except Exception as exc:
            self.show_error(f"Не удалось открыть Excel: {exc}")

    def show_error(self, message: str):
        self.statusBar().showMessage(message)
        QMessageBox.warning(self, "Внимание", message)

    def closeEvent(self, event):
        self.disconnect_serial()
        event.accept()


def main():
    app = QApplication(sys.argv)
    app.setApplicationName(APP_TITLE)
    app.setWindowIcon(load_app_icon())
    window = MainWindow()
    window.showMaximized()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
